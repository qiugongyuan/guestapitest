import requests
import openpyxl


# requests库写接口函数
def addevent(url, data):
    re = requests.post(url, data=data)
    response = re.json()
    return response


# 写入断言结果到excel里
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)  # 加载文件
    sheet = wb[sheetname]  # 读取sheet
    sheet.cell(row=row, column=column).value = final_result  # 将结果写入cell
    wb.save(filename)  # 保存excel


def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 获取表单
    max_raw = sheet.max_row  # 获取最大行数
    case_list = []  # 创建空列表,存放测试用例
    for i in range(2, max_raw + 1):
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,  # 获取case_id
            url=sheet.cell(row=i, column=5).value,  # 获取URL
            data=sheet.cell(row=i, column=6).value,  # 获取data传参
            expect=sheet.cell(row=i, column=7).value  # 获取期望结果
        )
        case_list.append(dict1)  # 每循环一次就将一行数据加入列表
    return case_list  # 返回测试用例列表


cases = read_data("D:\\qgy work\\useful\\pythondata.xlsx", "testdata")  # 调用读取excel文件函数
print(cases)
for case in cases:
    case_id = case.get("case_id")  # 从列表中取出case_id
    url = case.get("url")  # 从列表中取出url
    data = eval(case.get("data"))  # 从列表中取出data
    expect = eval(case.get("expect"))  # 从列表中取出expect
    expect_msg = expect.get("message")  # 从expect中取出message
    real_result = addevent(url=url, data=data)  # 调用添加事件接口函数将从列表中取出的url,data作为参数注册函数
    real_msg = real_result.get("message")
    print("预期结果的msg:{}".format(expect_msg))
    print("实际记过的msg:{}".format(real_msg))
    if real_msg == expect_msg:
        print("第{}条测试用例通过".format(case_id))
        final_re = "passed"
    else:
        print("第{}条测试用例不通过".format(case_id))
        final_re = "failed"
    write_result("D:\\qgy work\\useful\\pythondata.xlsx", "testdata", case_id + 1, 8, final_re)
